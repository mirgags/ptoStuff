#!/usr/bin/env
# -*- coding: utf-8 -*-

import json
from decimal import *
from openpyxl import load_workbook

class Attendance:
    def __init__(self):
        f = open("config.json")
        config = f.read()
        f.close()
        self.config = json.loads(config)
        self.filePath = self.config['filePath']

    def populate(self):
        sheetJson = openRead(self.filePath)
        self.data = sheetJson

    def balanceReport(self, ptoType, name):
        if ptoType == 'Sick':
            return self.data[name]['Sick']['Usable']
        else:
            return self.data[name][ptoType]


def openRead(filePath):
    wb = load_workbook(filename=filePath, data_only=True)
    ptoList = ['Vacation','Sick','Float']
    ptoDict = {}
    for ptoType in ptoList:
        ptoDict = readBalanceSheet(wb, ptoType, ptoDict)
    return ptoDict


def readBalanceSheet(workbook, sheetName, globalDict):
    c = 0
    ws = workbook[sheetName]
    inactive = False
    col = 'N'
    for row in ws.rows:
        c +=1
        if ws['B'+str(c)].value and ('INNACTIVE EMPLOYEES' in ws['B'+str(c)].value):
            inactive = True
        if (c >= 7 and ws['B'+str(c)].value) and not inactive:
            key = str(ws['B'+str(c)].value)
            if not key in globalDict:
                globalDict[key] = {}
            #print key + ':' + sheetName + ':' + str(ws[col+str(c)].value)
            #print len(str(ws['B'+str(c)].value))
            if sheetName == 'Sick':
                globalDict[key][sheetName] = {'Balance':0.0,'Usable':0.0}
                if ws[col+str(c)].value > 0:
                    globalDict[key][sheetName]['Balance']=round(ws[col+str(c)].value,2)
                else:
                    globalDict[key][sheetName]['Balance'] = 0.0
                if ws['R'+str(c)].value:
                    globalDict[key][sheetName]['Usable']=40 - round(ws['M'+str(c)].value,2)
                else:
                    globalDict[key][sheetName]['Usable'] = 40.0
            else:
                if sheetName == 'Float':
                    col = 'K'
                if ws[col+str(c)].value > 0:
                    #globalDict[key][sheetName] = Decimal(ws[col+str(c)].value)
                    globalDict[key][sheetName] = round(ws[col+str(c)].value, 2)
                else:
                    globalDict[key][sheetName] = 0.0
    #print 'count: ' + str(c)
    return globalDict


if __name__ == '__main__':
    getcontext().prec = 2
    obj = Attendance()
    obj.populate()
    rpt = obj.balanceReport('Vacation', 'McFarlane, Rob')
    print json.dumps('Vac: ' + str(rpt))
    rpt = obj.balanceReport('Sick', 'McFarlane, Rob')
    print json.dumps('Sick: ' + str(rpt))
    rpt = obj.balanceReport('Float', 'McFarlane, Rob')
    print json.dumps('Float: ' + str(rpt))
    path = '/Users/mmiraglia/Google Drive/Operations/Attendance/PTO Balances/pto.json'
    f = open(path, 'wb')
    f.write(json.dumps(obj.data, sort_keys=True, indent=2))
    f.close()
